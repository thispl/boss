
from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Oragadam Warehouse'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["","","","","Basic Salary","","","","","","","","","","","","Payment days & Extra days","","","","","","FIXED","","","","VARIABLES","","","","","","","","Fixed Earned","","","","Variables Earned","","","","","","","","","","Deductions","","","","","","","","","Employer's Part","","","","","",])
    ws.append(["S.No","Emp.No","","Name","Role","Role","DOB ","Gender","Project","Days","A/C NO","Bank","IFSC","PF Number","ESI Number","Service Charge","Working days","Paid Holiday","Performance day","Total days","Extra Hours","Food Deduction","Basic Wages","DA/Special Allowance","Leave with wages","Advance Bonus","HRA","Medical Reimburesement","Conveyance Allowance","MIP","Washing Allowance","Mobile Allowance","Other Allowance","Attendance Bonus","Gross Salary","Basic Salary","DA / Special Allowance","Leave with wages","Advance Bonus","HRA","Medical Reimburesement","Conveyance Allowance","MIP","Apr-20 Arrear HRA","Mobile Allowance","Other Allowance","Attendance Bonus","Extra Salary","Gross Salary","Employee Contribution PF","Employee Contributio ESI","Food","Professional Tax","LWF","Salary Advance","Safety shoe & Uniform","Total Deduction","Net pay","Employer Contribution PF","Employer Contribution ESI","LWF","Safety shoe & Uniform","Travel Allowance","CTC","Service Charges,","CTC + Service CHARGES","","IFSC Code","A/C No.","A/C Holder Name","","UAN No","ESIC No","Employee Contribution PF","Employee Contributiob ESI","Adhaar No.","Father name","DOB","Year","Age","Mobile no","Address","Location"])
    salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,'site':args.site},['*']) 

    i=1
    for ss in salary_slips:
        basic = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name },['amount']) or 0
        emp = frappe.get_doc('Employee',ss.employee,['*'])
        sc = frappe.get_value('Salary Detail',{'salary_component':'Service Charge','parent':ss.name},['amount']) or 0
        lww = frappe.get_value('Salary Detail',{'salary_component':'Leave With Wages','parent':ss.name},['amount']) or 0
        fd = frappe.get_value('Salary Detail',{'salary_component':'Food Deduction','parent':ss.name},['amount']) or 0
        basic = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name },['amount']) or 0
        da = frappe.get_value('Salary Detail',{'salary_component':'Dearness Allowance','parent':ss.name},['amount']) or 0
        ab = frappe.get_value('Salary Detail',{'salary_component':'Advance Bonus','parent':ss.name},['amount']) or 0
        hra = frappe.get_value('Salary Detail',{'salary_component':'House Rent Allowance','parent':ss.name},['amount']) or 0
        mr = frappe.get_value('Salary Detail',{'salary_component':'Medical Reimburesement','parent':ss.name},['amount']) or 0
        wa = frappe.get_value('Salary Detail',{'salary_component':'Washing Allowance','parent':ss.name},['amount']) or 0
        ma = frappe.get_value('Salary Detail',{'salary_component':'Mobile Allowance','parent':ss.name},['amount']) or 0
        oa = frappe.get_value('Salary Detail',{'salary_component':'Other Allowance','parent':ss.name},['amount']) or 0
        abb = frappe.get_value('Salary Detail',{'salary_component':'Attendance Bonus','parent':ss.name},['amount']) or 0
        pf = frappe.get_value('Salary Detail',{'salary_component':'Provident Fund','parent':ss.name},['amount']) or 0
        pt = frappe.get_value('Salary Detail',{'salary_component':'Professional Tax','parent':ss.name},['amount']) or 0
        esi = frappe.get_value('Salary Detail',{'salary_component':"Employees' State Insurance",'parent':ss.name},['amount']) or 0
        sa = frappe.get_value('Salary Detail',{'salary_component':'Salary Advance','parent':ss.name},['amount']) or 0
        su = frappe.get_value('Salary Detail',{'salary_component':'Safety shoe & Uniform','parent':ss.name},['amount']) or 0
        epf = frappe.get_value('Salary Detail',{'salary_component':'Employer PF','parent':ss.name},['amount']) or 0
        eesi = frappe.get_value('Salary Detail',{'salary_component':'Employer ESI','parent':ss.name},['amount']) or 0
        ta = frappe.get_value('Salary Detail',{'salary_component':'Travel Allowance','parent':ss.name},['amount']) or 0
        ctc = frappe.get_value('Salary Detail',{'salary_component':'Cost to Company','parent':ss.name},['amount']) or 0
        sc = frappe.get_value('Salary Detail',{'salary_component':'Service Charge','parent':ss.name},['amount']) or 0

        cs  = ctc + sc
    
        ws.append([i,ss.employee,"",ss.employee_name,"","","",emp.gender,ss.project,"","","","","","",emp.service_charge,ss.payment_days,"","",ss.total_working_days,"",fd,emp.basic,da,emp.leave_with_wages,emp.advance_bonus,emp.house_rent_allowance,mr,"","",wa,ma,oa,abb,ss.gross_pay,basic,da,lww,ab,emp.house_rent_allowance,mr,"","","",ma,oa,abb,"",ss.gross_pay,pf,esi,fd,pt,"",sa,su,ss.total_deduction,ss.net_pay,epf,eesi,"",su,ta,ctc,sc,cs,"","",ss.bank_account_no,ss.employee,"",ss.uan_number,ss.esi,pf,esi,"",emp.father_name,emp.date_of_birth,"","",emp.cell_number,emp.permanent_address,""])
        i=1+i
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=4)
    ws.merge_cells(start_row=1,start_column=5,end_row=1,end_column=12)
    ws.merge_cells(start_row=1,start_column=17,end_row=1,end_column=21)
    ws.merge_cells(start_row=1,start_column=23,end_row=1,end_column=26)
    ws.merge_cells(start_row=1,start_column=27,end_row=1,end_column=32)
    ws.merge_cells(start_row=1,start_column=35,end_row=1,end_column=38)
    ws.merge_cells(start_row=1,start_column=39,end_row=1,end_column=47)
    ws.merge_cells(start_row=1,start_column=49,end_row=1,end_column=57)
    ws.merge_cells(start_row=1,start_column=58,end_row=1,end_column=66)
    
    align_center = Alignment(horizontal='center')
    vertical = Alignment(textRotation=180,vertical='center',horizontal='center')

    for cell in ws["1:1"]:
        cell.alignment = align_center
    for cell in ws["2:2"]:
        cell.alignment = vertical

    bold_font = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = bold_font

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+2, min_col=23, max_col=26):
        for cell in rows:
            cell.fill = PatternFill(fgColor="f6fc2d", fill_type = "solid")

    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+2, min_col=27, max_col=34):
        for cell in rows:
            cell.fill = PatternFill(fgColor="f6fc2d", fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=49, max_col=49):
        for cell in rows:
            cell.fill = PatternFill(fgColor="b3d9ff", fill_type = "solid")
        
    for rows in ws.iter_rows(min_row=3, max_row=len(salary_slips)+2, min_col=49, max_col=49):
        for cell in rows:
            cell.fill = PatternFill(fgColor="d6d6c2", fill_type = "solid")

    for rows in ws.iter_rows(min_row=2, max_row=2, min_col=58, max_col=58):
        for cell in rows:
            cell.fill = PatternFill(fgColor="70db70", fill_type = "solid")
        
    for rows in ws.iter_rows(min_row=3, max_row=len(salary_slips)+2, min_col=58, max_col=58):
        for cell in rows:
            cell.fill = PatternFill(fgColor="99ccff", fill_type = "solid")
    
    for rows in ws.iter_rows(min_row=2, max_row=len(salary_slips)+2, min_col=64, max_col=64):
        for cell in rows:
            cell.fill = PatternFill(fgColor="99ccff", fill_type = "solid")
        
    for rows in ws.iter_rows(min_row=2, max_row=len(salary_slips)+2, min_col=65, max_col=65):
        for cell in rows:
            cell.fill = PatternFill(fgColor="99ccff", fill_type = "solid")

    

    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+2, min_col=1, max_col=len(salary_slips)+83):

        for cell in rows:
            cell.border = border


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'