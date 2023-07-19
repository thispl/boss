from __future__ import unicode_literals
from os import stat
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, touch_file
from frappe import _, log_error
from frappe.utils.csvutils import UnicodeWriter, build_csv_response, read_csv_content
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime, time

import openpyxl
from openpyxl import Workbook
import re
from openpyxl.styles import Font, Alignment,Border,Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


@frappe.whitelist()
def download():
    filename = 'dsv_register'
    test = build_xlsx_response(filename)


# return xlsx file object
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    header = title1(args)   
    ws.append(header)
    header = title2(args)   
    ws.append(header)
    header = title3(args)   
    ws.append(header)
    # header = title4(args)   
    # ws.append(header)
    header = title6(args)
    ws.append(header)
    header = title5(args)   
    ws.append(header)
    header = title7(args)   
    ws.append(header)
    header = title8(args)   
    ws.append(header)
    
     
    data = get_data(args)

    for row in data:
        ws.append(row)

    ws.sheet_view.zoomScale = 80

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    # write out response as a xlsx type
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'


@frappe.whitelist()
def title1(args):
    contractor_name = '-'
    if args.contractor :
        contractor_name = args.contractor
    data=['','','','','','','','','','FORM XXVII',]
    data.append(contractor_name)
    return data

@frappe.whitelist()
def title2(args):
    data=['','','','','','','','(See Rule 78(1) (a) (i) of the Tamil Nadu Contract Labour Rules,1975)',]
    return data

@frappe.whitelist()
def title3(args):
    data=['','','','','Wage Period: Monthly',]
    data.append(args.from_date)
    data.append('to')
    data.append(args.to_date)
   
    return data

# @frappe.whitelist()
# def title4(args):
#     data=['date',]
#     return data

@frappe.whitelist()
def title5(args):
    data = ['','','','','','','','','Register of Wages']
    return data

@frappe.whitelist()
def title6(args):
    data=['Name and Address of contractor :',args.company,'','',]
    data.append('Name and Address of establishment in /under which contract is carried on :',)
    data.append(args.report)
    return data

@frappe.whitelist()
def title7(args):
    data=['Name and Location of Work :' ]
    # data.append(args.client)
    return data

@frappe.whitelist()
def title8(args):
    data=['Sr no','Employee No','Employee Name','Role','DOJ','No of days Worked','Basic wages','DA','Leave with Wages','Advance Bonus','HRA','Other allowance','OT Amount','Overtime HRS','Total','P.F','E.S.I','Food','PT','Salary Advance','Safety Shoe & Uniform','Total Deduction','Travel allowance','Net Amount Paid','Sign or Thumb Impression of Workman',]
    return data

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates

@frappe.whitelist()
def get_data(args):
    data=[]
    values = frappe.db.sql("select name,`tabSalary Slip`.employee_name as employee_name,`tabSalary Slip`.employee as employee,`tabSalary Slip`.date_of_join as doj,`tabSalary Slip`.payment_days as payment_days,`tabSalary Slip`.rounded_total as rounded,`tabSalary Slip`.total_deduction as deduction,`tabSalary Slip`.net_pay as net_pay ,`tabSalary Slip`.total_working_hours as total_wh from `tabSalary Slip` where `tabSalary Slip`.start_date between '%s'and '%s'"%(args.from_date,args.to_date),as_dict=True)
    if args.company:
        values = frappe.db.sql("select name,`tabSalary Slip`.employee_name as employee_name,`tabSalary Slip`.employee as employee,`tabSalary Slip`.date_of_join as doj,`tabSalary Slip`.payment_days as payment_days,`tabSalary Slip`.rounded_total as rounded,`tabSalary Slip`.total_deduction as deduction,`tabSalary Slip`.net_pay as net_pay,`tabSalary Slip`.total_working_hours as total_wh from `tabSalary Slip` where `tabSalary Slip`.company = '%s' and `tabSalary Slip`.start_date between '%s'and '%s'"%(args.company,args.from_date,args.to_date),as_dict=True)
    i = 1
    for value in values:
        row = []
        row.extend([i,value.employee,value.employee_name,])
        basic = frappe.db.get_value("Salary Detail",{'abbr':'B','parent':value.name},'amount')
        da = frappe.db.get_value("Salary Detail",{'abbr':'DA','parent':value.name},'amount')
        house_all = frappe.db.get_value("Salary Detail",{'abbr':'HRA','parent':value.name},'amount')
        lww = frappe.db.get_value("Salary Detail",{'abbr':'LWW','parent':value.name},'amount')
        oa = frappe.db.get_value("Salary Detail",{'abbr':'OA','parent':value.name},'amount')
        overtime = frappe.db.get_value("Salary Detail",{'abbr':'OT','parent':value.name},'amount')
        att_bonus = frappe.db.get_value("Salary Detail",{'abbr':'AB','parent':value.name},'amount')
        wa = frappe.db.get_value("Salary Detail",{'abbr':'WA','parent':value.name},'amount')
        pf = frappe.db.get_value("Salary Detail",{'abbr':'PF','parent':value.name},'amount')
        esi = frappe.db.get_value("Salary Detail",{'abbr':'ESI','parent':value.name},'amount')
        Canteen = frappe.db.get_value("Salary Detail",{'abbr':'CC','parent':value.name},'amount')
        advance = frappe.db.get_value("Salary Detail",{'abbr':'AB_1','parent':value.name},'amount')
        sa = frappe.db.get_value("Salary Detail",{'abbr':'SA','parent':value.name},'amount')
        su = frappe.db.get_value("Salary Detail",{'abbr':'SU','parent':value.name},'amount')
        pt = frappe.db.get_value("Salary Detail",{'abbr':'PT','parent':value.name},'amount')
        bank_no = frappe.db.get_value("Employee",{'employee':value.employee},['bank_ac_no',])
        designation = frappe.db.get_value("Employee",{'employee':value.employee},['designation'])
        ta = frappe.db.get_value("Salary Detail",{'abbr':'TA','parent':value.name},'amount')
       
        row.append(designation)
        row.append(value.doj)
        row.append(value.payment_days)
        row.append(basic)
        row.append(da)
        row.append(lww)
        row.append(advance)
        row.append(house_all)
        row.append(oa)
        row.append(overtime)
        row.append(value.total_wh)
        row.append(value.rounded)
        row.append(pf)
        row.append(esi)
        row.append('food')
        row.append(pt)
        row.append(sa)
        row.append(su)   
        
        row.append(value.deduction)
        row.append(ta)
        row.append(value.net_pay)
        row.append(bank_no)     
        data.append(row)
        i +=1
        
    return data